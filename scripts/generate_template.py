"""
엑셀 일괄 업로드 템플릿 생성 스크립트.

실행:
    python scripts/generate_template.py

산출물: 보유마 일괄등록_템플릿.xlsx (프로젝트 루트에 생성됨)

⚠️ 주의: 안내 문구는 반드시 데이터 시트와 "별도 시트"에 둔다.
   같은 시트 안에(데이터 표 아래 등) 안내를 적으면, import_service.parse_excel()이
   pandas로 시트 전체를 읽을 때 그 문구까지 데이터 행으로 잘못 인식해서
   미리보기 화면에 가짜 오류 행이 잔뜩 뜨는 문제가 실제로 있었다 (실사용 검증 중 발견).
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

OUTPUT_PATH = Path(__file__).resolve().parent.parent / "보유마 일괄등록_템플릿.xlsx"

_HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="1C64F2", end_color="1C64F2", fill_type="solid")
_EXAMPLE_FILL = PatternFill(start_color="FFF9DB", end_color="FFF9DB", fill_type="solid")

_HEADERS = ["마명", "마종", "등록번호", "품종코드"]

# import_service.parse_excel()이 실제로 요구하는 컬럼(REQUIRED_COLS)과 반드시 일치해야 한다.
_EXAMPLE_ROWS = [
    ["닉스고", "씨수말", "0041819", "00100"],
    ["번개", "위수탁마", "0060446", "00100"],
]

_NOTES = [
    "이 시트는 참고용이며, 업로드 시 파싱되지 않습니다.",
    "",
    "- 마명, 마종, 등록번호는 필수 입력 항목입니다.",
    "- 마종은 씨수말/교육마/관상마/위수탁마/기타마 중 하나여야 합니다.",
    "- 등록번호, 품종코드는 horsepia.com 해당 말 URL의 hrNo, hrsGbCd 값을 그대로 넣으세요.",
    "- 품종코드는 비워두면 조회 시 자동으로 찾아 저장됩니다.",
    '- "보유마 일괄등록" 시트의 노란색 예시 행은 실제 데이터로 교체하거나 삭제한 뒤 업로드하세요.',
]


def build_template() -> Workbook:
    wb = Workbook()

    ws = wb.active
    ws.title = "보유마 일괄등록"
    for col, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(_EXAMPLE_ROWS, start=2):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = Font(name="맑은 고딕")
            cell.fill = _EXAMPLE_FILL

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12

    # 안내는 반드시 별도 시트에 (데이터 시트에 두면 파싱 오류 발생 — 위 모듈 docstring 참고)
    notes_ws = wb.create_sheet("안내")
    for i, note in enumerate(_NOTES, start=1):
        notes_ws.cell(row=i, column=1, value=note).font = Font(name="맑은 고딕", size=10)
    notes_ws.column_dimensions["A"].width = 70

    return wb


def main() -> None:
    wb = build_template()
    wb.save(OUTPUT_PATH)
    print(f"템플릿 생성 완료: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()