"""
위탁 연간·전체 기간 통계 리포트 집계.

entrustment(모집단) × career_summary(통산 성적) 조인.
경주 지표는 사업연도 출주에 한정하지 않고 작성 시점 통산을 사용한다.
"""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date
from typing import Any, Literal, Optional
import io

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from config.constants import STATUS_ENTRUSTED, STATUS_ENDED
from db import repository
from services import entrustment_service
from shared.applicant_name import normalize_applicant_name

ReportScope = Literal["year", "all"]
ReportLayout = Literal["status", "racing"]


@dataclass
class ReportKpi:
    label: str
    value: str
    unit: str = ""


@dataclass
class ApplicantRow:
    applicant_name: str
    cells: dict[str, str]


@dataclass
class HorseWinRow:
    name: str
    applicant_name: str
    starts: str
    wins: str
    prize_won: str


@dataclass
class ReportPreview:
    scope: ReportScope
    layout: ReportLayout
    title: str
    subtitle: str
    application_year: Optional[int]
    kpis: list[ReportKpi] = field(default_factory=list)
    applicant_headers: list[str] = field(default_factory=list)
    applicant_rows: list[ApplicantRow] = field(default_factory=list)
    win_horse_rows: list[HorseWinRow] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)
    is_stub: bool = False


def build_report_preview(
    scope: ReportScope,
    application_year: Optional[int] = None,
) -> ReportPreview:
    if scope == "year":
        if application_year is None:
            raise ValueError("사업연도별 집계에는 사업연도가 필요합니다.")
        horses = entrustment_service.search_horses(
            application_year=application_year
        )
        scope_label = f"사업연도 {application_year} 계약"
    else:
        horses = entrustment_service.list_horses()
        application_year = None
        scope_label = "전체 기간 위탁 계약 (마번 중복 제거 후 성적 합산)"

    career_map = _career_map()
    entrust_n = sum(1 for h in horses if h.status == STATUS_ENTRUSTED)
    layout: ReportLayout = (
        "racing" if scope == "all" or entrust_n == 0 else "status"
    )

    if layout == "status":
        assert application_year is not None
        return _build_status_year(
            year=application_year,
            horses=horses,
            scope_label=scope_label,
        )
    return _build_racing(
        scope=scope,
        year=application_year,
        horses=horses,
        career_map=career_map,
        scope_label=scope_label,
        dedupe_horses=(scope == "all"),
    )


def _career_map() -> dict[str, dict[str, Any]]:
    rows = repository.list_all_career_summaries_with_horse()
    out: dict[str, dict[str, Any]] = {}
    for row in rows:
        hid = row.get("horse_id")
        if hid and hid not in out:
            out[hid] = row
    return out


def _applicant_key(name: Optional[str]) -> str:
    return normalize_applicant_name(name)


def _to_int(value: Any) -> int:
    try:
        if value is None:
            return 0
        return int(value)
    except (TypeError, ValueError):
        return 0


def _fmt_num(n: int | float) -> str:
    if isinstance(n, float):
        text = f"{n:,.1f}".rstrip("0").rstrip(".")
        return text
    return f"{n:,}"


def _fmt_won(won: int | float | None) -> str:
    """원 단위, 천 단위 콤마."""
    return f"{_to_int(won):,}"


def _win_rate(wins: int, starts: int) -> str:
    if starts <= 0:
        return "—"
    return f"{(wins / starts) * 100:.1f}"


def _notes(scope_label: str) -> list[str]:
    return [
        "단위: 두수=두, 금액=원(천 단위 콤마), 승률=%.",
        f"위탁 모집단: {scope_label}.",
        "경주 지표는 해당 연도 출주에 한정하지 않으며, "
        "작성 시점 career_summary(통산)를 마번으로 조인해 집계합니다.",
    ]


def _build_status_year(
    *,
    year: int,
    horses: list,
    scope_label: str,
) -> ReportPreview:
    total = len(horses)
    n_in = sum(1 for h in horses if h.status == STATUS_ENTRUSTED)
    n_end = sum(1 for h in horses if h.status == STATUS_ENDED)
    fee_sum = sum(_to_int(h.entrustment_fee) for h in horses)
    applicants = {_applicant_key(h.applicant_name) for h in horses}

    # 신청인별 상태·비용
    buckets: dict[str, list] = defaultdict(list)
    for h in horses:
        buckets[_applicant_key(h.applicant_name)].append(h)

    headers = [
        "신청인",
        "두수(두)",
        "위탁중(두)",
        "종료(두)",
        "위탁비(원)",
        "낙찰(두)",
        "낙찰가(원)",
    ]
    rows: list[ApplicantRow] = []
    for name, items in sorted(
        buckets.items(), key=lambda kv: len(kv[1]), reverse=True
    ):
        fee = sum(_to_int(h.entrustment_fee) for h in items)
        rows.append(
            ApplicantRow(
                name,
                {
                    "신청인": name,
                    "두수(두)": _fmt_num(len(items)),
                    "위탁중(두)": _fmt_num(
                        sum(1 for h in items if h.status == STATUS_ENTRUSTED)
                    ),
                    "종료(두)": _fmt_num(
                        sum(1 for h in items if h.status == STATUS_ENDED)
                    ),
                    "위탁비(원)": _fmt_won(fee),
                    "낙찰(두)": "—",
                    "낙찰가(원)": "—",
                },
            )
        )

    return ReportPreview(
        scope="year",
        layout="status",
        title=f"{year} 사업연도 위수탁 통계",
        subtitle="진행 연도 · 위탁 상태 중심",
        application_year=year,
        kpis=[
            ReportKpi("위탁 두수", _fmt_num(total), "두"),
            ReportKpi("위탁중", _fmt_num(n_in), "두"),
            ReportKpi("위탁종료", _fmt_num(n_end), "두"),
            ReportKpi("위탁비 합계", _fmt_won(fee_sum), "원"),
            ReportKpi("신청인 수", _fmt_num(len(applicants)), "명"),
            ReportKpi(
                "평균 위탁비",
                _fmt_won(round(fee_sum / total)) if total else "0",
                "원",
            ),
        ],
        applicant_headers=headers,
        applicant_rows=rows,
        win_horse_rows=[],
        notes=_notes(scope_label)
        + [
            "당해(상태) 요약에는 경주 성적 KPI를 넣지 않습니다. "
            "성적은 위탁종료 연도·전체 기간 리포트에서 통산 기준으로 봅니다.",
            "낙찰 열은 이후 경매 집계 연결 예정입니다.",
        ],
        is_stub=False,
    )


def _count_with_career(horses: list, career_map: dict) -> int:
    """출전 기록이 1회 이상인 통산 요약만 '성적 보유'로 센다."""
    n = 0
    for h in horses:
        cs = career_map.get(h.horse_id or "")
        if not cs:
            continue
        if _to_int(cs.get("total_starts")) > 0:
            n += 1
    return n


def _unique_by_horse(horses: list) -> list:
    seen: set[str] = set()
    out = []
    for h in horses:
        hid = h.horse_id
        if not hid or hid in seen:
            continue
        seen.add(hid)
        out.append(h)
    return out


def _build_racing(
    *,
    scope: ReportScope,
    year: Optional[int],
    horses: list,
    career_map: dict[str, dict[str, Any]],
    scope_label: str,
    dedupe_horses: bool,
) -> ReportPreview:
    # 성적 합산용 말 목록 (전체 기간은 마번 중복 제거)
    race_horses = _unique_by_horse(horses) if dedupe_horses else list(horses)
    # 신청인별: 전체 기간은 계약 수·유니크 마번을 구분해 표시
    by_applicant_contracts: dict[str, list] = defaultdict(list)
    for h in horses:
        by_applicant_contracts[_applicant_key(h.applicant_name)].append(h)

    by_applicant_unique: dict[str, list] = defaultdict(list)
    seen_pair: set[tuple[str, str]] = set()
    for h in race_horses:
        key = _applicant_key(h.applicant_name)
        pair = (key, h.horse_id or "")
        if pair in seen_pair:
            continue
        seen_pair.add(pair)
        by_applicant_unique[key].append(h)

    starts_t = wins_t = prize_t = 0
    for h in race_horses:
        cs = career_map.get(h.horse_id or "", {})
        starts_t += _to_int(cs.get("total_starts"))
        wins_t += _to_int(cs.get("total_wins"))
        prize_t += _to_int(cs.get("total_prize_money"))

    fee_sum = sum(_to_int(h.entrustment_fee) for h in horses)
    n_in = sum(1 for h in horses if h.status == STATUS_ENTRUSTED)
    n_end = sum(1 for h in horses if h.status == STATUS_ENDED)

    head_col = "누적위탁(두)" if scope == "all" else "위탁두수(두)"
    headers = [
        "신청인",
        head_col,
        "총출전(회)",
        "1착(회)",
        "승률(%)",
        "경주상금(원)",
        "1착 대표마",
    ]

    applicant_rows: list[ApplicantRow] = []
    sort_keys = sorted(
        by_applicant_unique.keys(),
        key=lambda k: len(by_applicant_unique[k]),
        reverse=True,
    )
    for name in sort_keys:
        items = by_applicant_unique[name]
        head_count = (
            len({h.horse_id for h in by_applicant_contracts[name] if h.horse_id})
            if scope == "all"
            else len(by_applicant_contracts[name])
        )
        starts = wins = prize = 0
        winners: list[tuple[str, int]] = []
        for h in items:
            cs = career_map.get(h.horse_id or "", {})
            s = _to_int(cs.get("total_starts"))
            w = _to_int(cs.get("total_wins"))
            p = _to_int(cs.get("total_prize_money"))
            starts += s
            wins += w
            prize += p
            if w > 0:
                winners.append((h.name or h.horse_id or "-", w))
        winners.sort(key=lambda t: t[1], reverse=True)
        rep = ", ".join(n for n, _ in winners[:3]) if winners else "—"
        applicant_rows.append(
            ApplicantRow(
                name,
                {
                    "신청인": name,
                    head_col: _fmt_num(head_count),
                    "총출전(회)": _fmt_num(starts),
                    "1착(회)": _fmt_num(wins),
                    "승률(%)": _win_rate(wins, starts),
                    "경주상금(원)": _fmt_won(prize),
                    "1착 대표마": rep,
                },
            )
        )

    # 1착 기록 말
    win_rows_data: list[tuple[int, int, HorseWinRow]] = []
    for h in race_horses:
        cs = career_map.get(h.horse_id or "", {})
        w = _to_int(cs.get("total_wins"))
        if w <= 0:
            continue
        s = _to_int(cs.get("total_starts"))
        p = _to_int(cs.get("total_prize_money"))
        win_rows_data.append(
            (
                w,
                p,
                HorseWinRow(
                    name=h.name or h.horse_id or "-",
                    applicant_name=_applicant_key(h.applicant_name),
                    starts=_fmt_num(s),
                    wins=_fmt_num(w),
                    prize_won=_fmt_won(p),
                ),
            )
        )
    win_rows_data.sort(key=lambda t: (t[0], t[1]), reverse=True)
    win_horse_rows = [t[2] for t in win_rows_data[:30]]

    if scope == "all":
        title = "위탁사업 전체 기간 누적 통계"
        subtitle = "누적 · 신청인·통산 성적 중심"
        kpis = [
            ReportKpi("누적 위탁 두수", _fmt_num(len(race_horses)), "두"),
            ReportKpi("현재 위탁중", _fmt_num(n_in), "두"),
            ReportKpi("누적 종료", _fmt_num(n_end), "두"),
            ReportKpi("위탁비 합계", _fmt_won(fee_sum), "원"),
            ReportKpi("누적 출전", _fmt_num(starts_t), "회"),
            ReportKpi("누적 1착", _fmt_num(wins_t), "회"),
            ReportKpi("승률 (1착÷출전)", _win_rate(wins_t, starts_t), "%"),
            ReportKpi(
                "경주상금 합계", _fmt_won(prize_t), "원"
            ),
        ]
    else:
        title = f"{year} 사업연도 위수탁 통계"
        subtitle = "종료 연도 · 경주 성과 중심"
        kpis = [
            ReportKpi("위탁 두수", _fmt_num(len(horses)), "두"),
            ReportKpi("위탁종료", _fmt_num(n_end), "두"),
            ReportKpi("위탁비 합계", _fmt_won(fee_sum), "원"),
            ReportKpi("총 출전 수", _fmt_num(starts_t), "회"),
            ReportKpi("1착 합계", _fmt_num(wins_t), "회"),
            ReportKpi("승률 (1착÷출전)", _win_rate(wins_t, starts_t), "%"),
            ReportKpi(
                "경주상금 합계", _fmt_won(prize_t), "원"
            ),
            ReportKpi(
                "성적 보유",
                _fmt_num(_count_with_career(race_horses, career_map)),
                "두",
            ),
        ]

    return ReportPreview(
        scope=scope,
        layout="racing",
        title=title,
        subtitle=subtitle,
        application_year=year,
        kpis=kpis,
        applicant_headers=headers,
        applicant_rows=applicant_rows,
        win_horse_rows=win_horse_rows,
        notes=_notes(scope_label),
        is_stub=False,
    )


def export_report_excel(
    scope: ReportScope,
    application_year: Optional[int] = None,
) -> bytes:
    """
    통계 미리보기와 동일 집계를 시트 1장 엑셀로 만든다.
    상단 KPI → 신청인 표 → (있으면) 1착 기록 말 → 집계 기준 주석.
    """
    preview = build_report_preview(scope, application_year)
    return _preview_to_excel_bytes(preview)


def report_excel_filename(
    scope: ReportScope,
    application_year: Optional[int] = None,
) -> str:
    if scope == "all":
        return "위탁통계_전체기간.xlsx"
    return f"위탁통계_{application_year}.xlsx"


def _preview_to_excel_bytes(preview: ReportPreview) -> bytes:
    wb = Workbook()
    ws = wb.active
    title = preview.title.replace("/", "-")[:31] or "위탁통계"
    ws.title = title

    ws.append([preview.title])
    ws.append([preview.subtitle])
    ws.append([])

    ws.append(["주요 지표", "값", "단위"])
    for kpi in preview.kpis:
        ws.append([kpi.label, kpi.value, kpi.unit])
    ws.append([])

    section = (
        "신청인별 경주 성과"
        if preview.layout == "racing"
        else "신청인별 위탁 집계"
    )
    ws.append([section])
    if preview.applicant_headers:
        ws.append(list(preview.applicant_headers))
        for row in preview.applicant_rows:
            ws.append(
                [row.cells.get(h, "") for h in preview.applicant_headers]
            )
    ws.append([])

    if preview.win_horse_rows:
        ws.append(["1착 기록 말"])
        win_headers = ["마명", "신청인", "출전(회)", "1착(회)", "상금(원)"]
        ws.append(win_headers)
        for h in preview.win_horse_rows:
            ws.append(
                [h.name, h.applicant_name, h.starts, h.wins, h.prize_won]
            )
        ws.append([])

    ws.append(["집계 기준"])
    for note in preview.notes:
        ws.append([note])

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
