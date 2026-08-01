"""
Phase 3: 경주성적 수집 오케스트레이션.
racing_scraper(순수 파싱)의 결과를 받아 RaceRecord/CareerSummary 테이블에 반영한다.
규칙: ui -> services(이 파일) -> repository 순서만 호출한다.

실행 정책 (설계문서 7.8절):
- 자동 스케줄러 없음. "경주기록 확인" 버튼 하나로만 트리거.
- 스크래핑 대상은 entrustment.status == '위탁종료' 이면서 아직 경주기록을 확인하지 않은 말
  (career_summary가 없거나 last_scraped_at이 비어있는 경우).
- 요청 간 딜레이를 두어 상대 서버 부하를 줄인다.

[통합 시 변경사항]
- repository.list_horses(status=...) 대상이 entrustment 테이블로 변경됨(기존 로직 그대로 재사용)
- last_scraped_at 파싱을 fromisoformat() 대신 방어적 _coerce_datetime()으로 변경
  (PostgreSQL TIMESTAMP 컬럼이라 psycopg가 이미 datetime 객체로 반환하는 경우 대응)
- get_race_records_with_horse_name() 신규 추가: race_record.horse_id를
  A의 horses.마번과 조회 시점에 JOIN해서 마명을 붙여 반환 (경주마명 표시 기능)
"""
from __future__ import annotations

import time
from dataclasses import dataclass, field
from datetime import datetime, timedelta

from adapters import racing_scraper
from adapters.racing_scraper import ScrapingError
from config.constants import STATUS_ENDED
from db import repository
from repository.horse_repository import _get_connection as _a_get_connection
from psycopg.rows import dict_row
from typing import Any

REQUEST_DELAY_SECONDS = 0.7
RECHECK_INTERVAL_DAYS = 7  # 위탁종료된 말은 이 기간마다 경주기록을 다시 확인한다


@dataclass
class RefreshResult:
    total_targets: int = 0
    success_count: int = 0
    no_race_history_count: int = 0
    failed_count: int = 0
    failed_details: list[str] = field(default_factory=list)


def _coerce_datetime(value):
    """last_scraped_at이 datetime 객체(정상)이거나 문자열(레거시)인 경우 모두 방어적으로 처리."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    try:
        return datetime.fromisoformat(value)
    except (ValueError, TypeError):
        return None


def _list_unverified_ended_horses() -> list[dict]:
    """
    status == '위탁종료' 인 말 중 아래 조건을 만족하는 말을 재확인 대상으로 삼는다.
    - career_summary가 없거나 last_scraped_at이 비어있는 경우 (한 번도 확인 안 함)
    - last_scraped_at이 RECHECK_INTERVAL_DAYS일보다 오래된 경우 (재확인 주기 도래)
    """
    ended = repository.list_horses(status=STATUS_ENDED)
    cutoff = datetime.now() - timedelta(days=RECHECK_INTERVAL_DAYS)

    unverified = []
    for horse in ended:
        summary = repository.get_career_summary(horse["horse_id"])
        if summary is None or summary.get("last_scraped_at") is None:
            unverified.append(horse)
            continue

        last_scraped = _coerce_datetime(summary["last_scraped_at"])
        if last_scraped is None or last_scraped < cutoff:
            unverified.append(horse)

    return unverified


def refresh_all_racehorses(
    delay_seconds: float = REQUEST_DELAY_SECONDS,
    progress_callback=None,
) -> RefreshResult:
    """
    위탁종료 상태이면서 아직 경주기록을 확인하지 않은 모든 말의 경주성적을 새로고침한다.
    말마다 기존 RaceRecord를 지우고 새로 채우는 전체 교체 방식(idempotent).
    """
    targets = _list_unverified_ended_horses()
    result = RefreshResult(total_targets=len(targets))
    session = racing_scraper.create_session()

    for i, horse in enumerate(targets):
        horse_id = horse["horse_id"]
        print(f"[{i+1}/{len(targets)}] 마번 {horse_id} 처리 중...")

        if progress_callback:
            progress_callback(i + 1, len(targets), horse_id)

        try:
            scraped = racing_scraper.fetch_race_data(horse_id, session=session)
        except ScrapingError as e:
            print(f"  실패: {e}")
            result.failed_count += 1
            result.failed_details.append(f"마번 {horse_id}: {e}")
            continue

        repository.delete_race_records_by_horse(horse_id)
        for record in scraped.race_records:
            repository.insert_race_record(record)

        if scraped.career_summary:
            summary = dict(scraped.career_summary)
            summary["last_scraped_at"] = datetime.now()
            repository.upsert_career_summary(summary)
            result.success_count += 1
        else:
            result.no_race_history_count += 1
            repository.upsert_career_summary(
                {
                    "horse_id": horse_id,
                    "total_starts": 0,
                    "total_wins": 0,
                    "win_rate": 0.0,
                    "total_prize_money": 0,
                    "rating": None,
                    "data_source": "scraping",
                    "last_scraped_at": datetime.now(),
                }
            )

        if i < len(targets) - 1:
            time.sleep(delay_seconds)

    return result


def get_race_records(horse_id: str) -> list[dict]:
    return repository.list_race_records(horse_id)


def get_career_summary(horse_id: str) -> dict | None:
    return repository.get_career_summary(horse_id)


def list_all_career_summaries() -> list[dict]:
    return repository.list_all_career_summaries_with_horse()


def get_race_records_with_horse_name(horse_id: str) -> list[dict]:
    """
    [신규] 경주성적 화면용. race_record를 A의 horses.마번과 조회 시점에 JOIN해서
    마명을 붙여 반환한다. race_record 테이블 자체에는 마명을 저장하지 않는다
    (스크래핑 응답에 마명 필드가 없고, A의 horses가 마명의 유일한 출처이므로).
    """
    with _a_get_connection() as conn:
        with conn.cursor(row_factory=dict_row) as cur:
            cur.execute(
                """
                SELECT rr.*, h.마명 AS horse_name
                FROM race_record rr
                JOIN horses h ON h.마번 = rr.horse_id
                WHERE rr.horse_id = %s
                ORDER BY rr.race_date DESC
                """,
                (horse_id,),
            )
            return [dict(r) for r in cur.fetchall()]

def get_race_stats_for_horse_ids(horse_ids: set[str]) -> dict[str, Any]:
    """주어진 마번 집합에 한정된 1위 두수 합계 / 전체 상금 합계."""
    summaries = repository.list_all_career_summaries_with_horse()
    filtered = [s for s in summaries if s["horse_id"] in horse_ids]
    return {
        "total_race_wins": sum(s.get("total_wins") or 0 for s in filtered),
        "total_prize_money": sum(s.get("total_prize_money") or 0 for s in filtered),
    }

import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

_EXPORT_COLUMNS = [
    ("마번", "horse_id"),
    ("마명", "horse_name"),
    ("위탁자", "applicant_name"),
    ("출주", "total_starts"),
    ("1위", "total_wins"),
    ("승률(%)", "win_rate"),
    ("총상금", "total_prize_money"),
    ("최종확인일", "last_scraped_at"),
]


def _format_scraped_date(value) -> str:
    if value is None:
        return "-"
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)[:10]


def export_career_summary_excel() -> bytes:
    """
    전체 통산성적 요약(racing_page.py 표와 동일 데이터/정렬)을 엑셀 바이트로 생성한다.
    이 엑셀은 조회 전용 출력물이며, 엑셀 일괄 등록(import_service.py)의 입력 템플릿과는
    무관하다 — 다시 업로드해서 반영하는 용도가 아니다.
    """
    summaries = list_all_career_summaries()
    summaries = sorted(summaries, key=lambda s: s.get("horse_name") or "")

    wb = Workbook()
    ws = wb.active
    ws.title = "통산성적요약"

    headers = [label for label, _ in _EXPORT_COLUMNS]
    ws.append(headers)

    for s in summaries:
        row = []
        for label, key in _EXPORT_COLUMNS:
            value = s.get(key)
            if key == "last_scraped_at":
                value = _format_scraped_date(value)
            row.append(value if value is not None else "")
        ws.append(row)

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()