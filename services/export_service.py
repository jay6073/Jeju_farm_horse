"""
조회 화면(main_page)에 표시된 마적사항을 A4 용지 크기에 맞춘 엑셀로 내보내는 서비스.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.worksheet import Worksheet

from models.horse import Horse

_KOREAN_FONT = "맑은 고딕"
_HEADER_FILL = PatternFill(start_color="1C64F2", end_color="1C64F2", fill_type="solid")


def build_horse_detail_excel(horse: Horse, basic_info: dict[str, str]) -> bytes:
    """
    말 한 마리의 마적사항(개체이력 기본정보)을 A4 세로 한 페이지에 맞춘
    엑셀 파일로 만들어 바이트로 반환한다. main_page의 "엑셀로 저장" 버튼에서 호출한다.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "마적사항"

    _configure_a4_page(ws)
    _write_title(ws, horse)
    _write_table(ws, basic_info)
    _autosize_columns(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _configure_a4_page(ws: Worksheet) -> None:
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.7, right=0.7, top=0.8, bottom=0.8, header=0.3, footer=0.3
    )
    ws.print_options.horizontalCentered = True


def _write_title(ws: Worksheet, horse: Horse) -> None:
    ws["A1"] = horse.마명
    ws["A1"].font = Font(name=_KOREAN_FONT, size=16, bold=True)
    ws.merge_cells("A1:B1")

    ws["A2"] = f"마종: {horse.마종}"
    ws["A2"].font = Font(name=_KOREAN_FONT, size=10, color="666666")
    ws.merge_cells("A2:B2")
    ws.append([])  # 빈 줄


def _write_table(ws: Worksheet, basic_info: dict[str, str]) -> None:
    start_row = 4
    header_row = ws[start_row]
    ws.cell(row=start_row, column=1, value="항목")
    ws.cell(row=start_row, column=2, value="내용")
    for col in (1, 2):
        cell = ws.cell(row=start_row, column=col)
        cell.font = Font(name=_KOREAN_FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, (label, value) in enumerate(basic_info.items(), start=start_row + 1):
        label_cell = ws.cell(row=i, column=1, value=label)
        value_cell = ws.cell(row=i, column=2, value=value)
        label_cell.font = Font(name=_KOREAN_FONT, size=10)
        value_cell.font = Font(name=_KOREAN_FONT, size=10)
        label_cell.alignment = Alignment(vertical="center")
        value_cell.alignment = Alignment(vertical="center")
        if (i - start_row) % 2 == 0:
            fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
            label_cell.fill = fill
            value_cell.fill = fill


def _autosize_columns(ws: Worksheet) -> None:
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40