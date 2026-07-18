"""엑셀 내보내기(A4 규격) 검증."""
import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import load_workbook

from models.horse import Horse
from services.export_service import build_horse_detail_excel


def check(label, condition):
    status = "OK " if condition else "FAIL"
    print(f"[{status}] {label}")
    assert condition, label


def main():
    horse = Horse(마명="닉스고", 마종="씨수말", 마번="0041819", 품종코드="00100")
    basic_info = {
        "마명부여일": "2019-01-26",
        "출생일": "2016-01-29 (10세)",
        "성별": "수",
        "품종": "더러브렛",
        "부마명": "PAYNTER",
        "모마명": "KOSMO'S BUDDY",
    }

    excel_bytes = build_horse_detail_excel(horse, basic_info)
    check("바이트가 비어있지 않음", len(excel_bytes) > 0)

    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active

    check("A4 용지로 설정됨", str(ws.page_setup.paperSize) == str(ws.PAPERSIZE_A4))
    check("세로 방향으로 설정됨", ws.page_setup.orientation == "portrait")
    check("한 페이지 너비에 맞춤 설정됨", ws.page_setup.fitToWidth == 1)
    check("한 페이지 높이에 맞춤 설정됨", ws.page_setup.fitToHeight == 1)
    check("fitToPage 활성화됨", ws.sheet_properties.pageSetUpPr.fitToPage is True)

    check("제목(마명)이 A1에 기록됨", ws["A1"].value == "닉스고")
    check("마종 정보가 A2에 포함됨", "씨수말" in ws["A2"].value)

    # 헤더(항목/내용) + 6개 데이터 행 = start_row(4) 기준 검증
    check("헤더 행 '항목' 존재", ws.cell(row=4, column=1).value == "항목")
    check("헤더 행 '내용' 존재", ws.cell(row=4, column=2).value == "내용")
    check("첫 데이터 행: 마명부여일", ws.cell(row=5, column=1).value == "마명부여일")
    check("첫 데이터 행 값: 2019-01-26", ws.cell(row=5, column=2).value == "2019-01-26")
    check(
        "마지막 데이터 행: 모마명",
        ws.cell(row=4 + len(basic_info), column=1).value == "모마명",
    )

    print("\n모든 검증 통과 ✅")


if __name__ == "__main__":
    main()